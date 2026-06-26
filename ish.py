# Sizning barcha talablaringizni inobatga olgan holda "📌 Eslatmalar" funksiyasini aynan siz xohlagandek (rasmdagidek) mukammal qilib sozladim.

# Qanday ishlaydi:

# Yangi eslatma qo'shganingizda u ro'yxatda ◻️ (bajarilmagan) holatda turadi.

# Ishni bajarib, uning ustiga bossangiz, u ✅ (bajarilgan) holatga o'tadi va ro'yxatning eng pastiga tushadi. Bajarilmagan ishlar o'z joyida, tepada qolaveradi.

# Agar bajarilgan (✅) ishning ustiga yana bir marta bossangiz, u ro'yxatdan butunlay o'chib ketadi (tozalab yuboriladi). Shunday qilib, ro'yxatni doim toza saqlashingiz mumkin.

# Mana to'liq, tayyor kod. Buni nusxalab botingizga joylashingiz mumkin:

# Python
# ================== 1. IMPORTLAR ==================
import os
import math
import random
import sqlite3
import threading
import re
import logging
import asyncio
import httpx
from datetime import datetime, timedelta, date
from http.server import SimpleHTTPRequestHandler, HTTPServer

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
try:
    from docx import Document
except ImportError:
    Document = None

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    KeyboardButton,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    ConversationHandler,
    filters,
    Application
)

# ================== 2. SOZLAMALAR ==================
TOKEN = "8802766629:AAHdLYyvunZlN4U2xzv4YL55eM6nKrMdakk"
BOSS_ID = 5247098284

WORK_START = "08:00"
WORK_END = "18:00"
LUNCH_START = "12:00"
LUNCH_END = "13:00"
WORK_HOURS_PER_DAY = 9  # 08:00–18:00 minus 1 soat tushlik = 9 soat

UZBEKISTAN_HOLIDAYS = {
    "01-01", "01-14", "03-08", "03-21", "03-22", "03-23",
    "05-09", "09-01", "10-01", "12-08",
}

def is_work_day(d: date) -> bool:
    if d.weekday() >= 5:
        return False
    if d.strftime("%m-%d") in UZBEKISTAN_HOLIDAYS:
        return False
    return True

def count_work_days_in_month(month_str: str) -> int:
    start = datetime.strptime(month_str + "-01", "%Y-%m-%d").date()
    end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    total = 0
    cur = start
    while cur <= end:
        if is_work_day(cur):
            total += 1
        cur += timedelta(days=1)
    return total

def calculate_late_fine(passport: str, arrived_time_str: str, daily_salary: int, month_str: str) -> dict:
    work_days = count_work_days_in_month(month_str)
    if work_days == 0:
        return {"late_minutes": 0, "fine": 0, "hourly_rate": 0}
    hourly_rate = daily_salary / WORK_HOURS_PER_DAY / work_days
    today = datetime.now()
    work_start_dt = today.replace(hour=8, minute=0, second=0, microsecond=0)
    arrived_dt = datetime.strptime(arrived_time_str, "%H:%M").replace(
        year=today.year, month=today.month, day=today.day
    )
    late_minutes = max(int((arrived_dt - work_start_dt).total_seconds() // 60), 0)
    fine = int(late_minutes * (hourly_rate / 60))
    return {"late_minutes": late_minutes, "fine": fine, "hourly_rate": hourly_rate}

# ================== 3. CONVERSATION STATES ==================
FULLNAME, PASSPORT, JSHSHIR, MAOSH, PIN_ADD, LOCATION = range(6)
KELDIM_CODE, PIN_STEP, BOSS_LOCATION = range(6, 9)
ISHBERISH_SELECT, ISHBERISH_MESSAGE = range(9, 11)
TOPSHIRIQ_JAVOB = 11
YANGI_ESLATMA = 12

# ================== 4. DATABASE ==================
conn = sqlite3.connect("ishchilar.db", check_same_thread=False)
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS ishchilar (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    fullname TEXT,
    passport TEXT UNIQUE,
    jshshir TEXT DEFAULT '',
    birthdate TEXT DEFAULT '',
    code INTEGER UNIQUE,
    role TEXT DEFAULT 'worker',
    daily_salary INTEGER DEFAULT 100000,
    pin TEXT DEFAULT '0000',
    password TEXT DEFAULT '',
    tg_id TEXT DEFAULT NULL,
    tax_percent INTEGER DEFAULT 12,
    location TEXT DEFAULT ''
)
""")

cursor.execute("PRAGMA table_info(ishchilar)")
existing_cols = [info[1] for info in cursor.fetchall()]
for col_name, col_type in [
    ("location", "TEXT DEFAULT ''"),
    ("birthdate", "TEXT DEFAULT ''"),
    ("password", "TEXT DEFAULT ''"),
    ("jshshir", "TEXT DEFAULT ''"),
]:
    if col_name not in existing_cols:
        cursor.execute(f"ALTER TABLE ishchilar ADD COLUMN {col_name} {col_type}")
        conn.commit()

cursor.execute("""
CREATE TABLE IF NOT EXISTS davomat (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    passport TEXT,
    date TEXT,
    time TEXT,
    confirmed INTEGER DEFAULT 0
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS moliya (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    passport TEXT,
    date TEXT,
    late_minutes INTEGER DEFAULT 0,
    jarima INTEGER DEFAULT 0,
    avans INTEGER DEFAULT 0
)
""")

# Yangilangan messages jadvali
cursor.execute("""
CREATE TABLE IF NOT EXISTS messages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    thread_id TEXT DEFAULT NULL,
    ishchi_id INTEGER,
    passport TEXT,
    sender TEXT,
    type TEXT,
    content TEXT,
    date TEXT,
    is_read INTEGER DEFAULT 0
)
""")

# Eslatmalar uchun yangi jadval
cursor.execute("""
CREATE TABLE IF NOT EXISTS eslatmalar (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    boss_id INTEGER,
    text TEXT,
    is_done INTEGER DEFAULT 0,
    created_at TEXT
)
""")

cursor.execute("PRAGMA table_info(messages)")
msg_cols = [info[1] for info in cursor.fetchall()]
if "thread_id" not in msg_cols:
    cursor.execute("ALTER TABLE messages ADD COLUMN thread_id TEXT DEFAULT NULL")
    conn.commit()

cursor.execute("""
CREATE TABLE IF NOT EXISTS boss_location (
    id INTEGER PRIMARY KEY,
    map_link TEXT
)
""")

cursor.execute("""
    INSERT INTO boss_location (id, map_link)
    VALUES (1, ?)
    ON CONFLICT(id) DO UPDATE SET map_link=excluded.map_link
""", ("40.861751,71.458907",))
conn.commit()

cutoff = (datetime.now() - timedelta(days=60)).strftime("%Y-%m-%d")
cursor.execute("DELETE FROM davomat WHERE date < ?", (cutoff,))
conn.commit()

# ================== 5. YORDAMCHI FUNKSIYALAR ==================
def generate_code():
    while True:
        c = random.randint(100000, 999999)
        cursor.execute("SELECT 1 FROM ishchilar WHERE code=?", (c,))
        if not cursor.fetchone():
            return c

def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371
    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)
    a = math.sin(d_lat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(d_lon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def parse_latlon_from_link(link):
    match = re.search(r'q=([0-9\.\-]+),([0-9\.\-]+)', link)
    if match:
        return float(match.group(1)), float(match.group(2))
    coords = re.search(r'@([0-9\.\-]+),([0-9\.\-]+)', link)
    if coords:
        return float(coords.group(1)), float(coords.group(2))
    return None, None

def get_month_days_and_attendance(month=None):
    if not month:
        month = datetime.now().strftime("%Y-%m")
    conn_local = sqlite3.connect("ishchilar.db")
    cur = conn_local.cursor()
    cur.execute("SELECT fullname, passport, daily_salary FROM ishchilar WHERE role='worker'")
    workers = cur.fetchall()
    start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
    end_date = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    days = []
    cur_d = start_date
    while cur_d <= end_date:
        days.append(cur_d.strftime("%Y-%m-%d"))
        cur_d += timedelta(days=1)
    result = []
    for fullname, passport, salary in workers:
        cur.execute("SELECT date FROM davomat WHERE passport=? AND confirmed=1 AND date LIKE ?", (passport, f"{month}%"))
        kelgan_sanalar = {r[0] for r in cur.fetchall()}
        result.append((fullname, passport, salary, kelgan_sanalar))
    conn_local.close()
    return days, result

def oylik_maosh_hisob(passport, month_str=None):
    if not month_str:
        month_str = datetime.now().strftime("%Y-%m")
    ish_kunlari = count_work_days_in_month(month_str)
    conn_local = sqlite3.connect("ishchilar.db")
    cur = conn_local.cursor()
    cur.execute("SELECT fullname, daily_salary FROM ishchilar WHERE passport=?", (passport,))
    row = cur.fetchone()
    if not row:
        conn_local.close()
        return None
    fullname, salary = row
    cur.execute("SELECT COUNT(*) FROM davomat WHERE passport=? AND confirmed=1 AND date LIKE ?", (passport, f"{month_str}%"))
    kelgan = cur.fetchone()[0] or 0
    cur.execute("SELECT SUM(late_minutes), SUM(jarima) FROM moliya WHERE passport=? AND date LIKE ?", (passport, f"{month_str}%"))
    moliya_data = cur.fetchone()
    total_late_minutes = moliya_data[0] or 0
    total_jarima = moliya_data[1] or 0
    cur.execute("SELECT SUM(avans) FROM moliya WHERE passport=? AND date LIKE ?", (passport, f"{month_str}%"))
    total_avans = cur.fetchone()[0] or 0
    conn_local.close()
    kelmagan = max(ish_kunlari - kelgan, 0)
    bir_kunlik = salary / ish_kunlari if ish_kunlari else 0
    yoqchilik = kelmagan * bir_kunlik
    
    sof_maosh = salary - yoqchilik - total_avans
    
    return {
        "fullname": fullname,
        "oylik": salary,
        "ish_kunlari": ish_kunlari,
        "kelgan": kelgan,
        "kelmagan": kelmagan,
        "late_minutes": total_late_minutes,
        "bir_kunlik": int(bir_kunlik),
        "yoqchilik": int(yoqchilik),
        "jarima": int(total_jarima),
        "avans": int(total_avans),
        "sof_maosh": int(sof_maosh),
    }

def export_oylik_maosh_excel_for(passport, month=None):
    month_str = month if month else datetime.now().strftime("%Y-%m")
    hisob = oylik_maosh_hisob(passport, month_str)
    if not hisob:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Shaxsiy Oylik Maosh"
    
    HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    DATA_FILL = PatternFill("solid", start_color="E9F1F8", end_color="E9F1F8")
    RESULT_FILL = PatternFill("solid", start_color="C6E0B4", end_color="C6E0B4")
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"Oylik Maosh Hisoboti - {month_str}"
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = HEADER_FILL
    t.alignment = CENTER

    data = [
        ("F.I.Sh:", hisob["fullname"]),
        ("Ish kunlari (oy):", f"{hisob['ish_kunlari']} kun"),
        ("Kelgan kunlar:", f"{hisob['kelgan']} kun"),
        ("Kelmagan kunlar:", f"{hisob['kelmagan']} kun"),
        ("Kechikish (daqiqa):", f"{hisob['late_minutes']} daqiqa"),
        ("Jarima summasi:", f"{hisob['jarima']:,} so'm"),
        ("Avans:", f"{hisob['avans']:,} so'm"),
        ("SOF MAOSH:", f"{hisob['sof_maosh']:,} so'm")
    ]

    for row_idx, (col1, col2) in enumerate(data, start=2):
        c1 = ws.cell(row=row_idx, column=1, value=col1)
        c1.font = Font(bold=True)
        c1.border = BORDER
        c1.fill = DATA_FILL
        
        c2 = ws.cell(row=row_idx, column=2, value=col2)
        c2.alignment = LEFT
        c2.border = BORDER
        
        if col1 == "SOF MAOSH:":
            c1.fill = RESULT_FILL
            c2.fill = RESULT_FILL
            c2.font = Font(bold=True, color="006100")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30

    file_name = f"oylik_maosh_{passport}_{month_str}.xlsx"
    wb.save(file_name)
    return file_name

def export_worker_davomat_excel(passport, month=None):
    if not month:
        month = datetime.now().strftime("%Y-%m")
    start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
    end_date = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    days = []
    cur_d = start_date
    while cur_d <= end_date:
        days.append(cur_d.strftime("%Y-%m-%d"))
        cur_d += timedelta(days=1)
    conn_local = sqlite3.connect("ishchilar.db")
    cur = conn_local.cursor()
    cur.execute("SELECT fullname FROM ishchilar WHERE passport=?", (passport,))
    row = cur.fetchone()
    if not row:
        conn_local.close()
        return None
    fullname = row[0]
    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"
    headers = ["Ism Familiya"]
    for d in days:
        d_obj = datetime.strptime(d, "%Y-%m-%d").date()
        marker = d[-2:]
        if not is_work_day(d_obj):
            marker += "(dam)"
        headers.append(marker)
    headers += ["Kelgan", "Kelmagan (ish kuni)"]
    ws.append(headers)
    cur.execute("SELECT date FROM davomat WHERE passport=? AND confirmed=1 AND date LIKE ?", (passport, f"{month}%"))
    kelgan = {r[0] for r in cur.fetchall()}
    row_data = [fullname]
    for d in days:
        d_obj = datetime.strptime(d, "%Y-%m-%d").date()
        if not is_work_day(d_obj):
            row_data.append("🔵")
        elif d in kelgan:
            row_data.append("✅")
        else:
            row_data.append("❌")
    ish_kunlari = count_work_days_in_month(month)
    kelgan_ish = len([d for d in days if d in kelgan and is_work_day(datetime.strptime(d, "%Y-%m-%d").date())])
    kelmagan_cnt = max(ish_kunlari - kelgan_ish, 0)
    row_data += [kelgan_ish, kelmagan_cnt]
    ws.append(row_data)
    file_name = f"davomat_{passport}_{month}.xlsx"
    wb.save(file_name)
    conn_local.close()
    return file_name

def update_excel_report(month=None):
    if not month:
        month = datetime.now().strftime("%Y-%m")
    GREEN_FILL  = PatternFill("solid", start_color="92D050", end_color="92D050")
    RED_FILL    = PatternFill("solid", start_color="FF6666", end_color="FF6666")
    YELLOW_FILL = PatternFill("solid", start_color="FFD700", end_color="FFD700")
    HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    NAME_FILL   = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    TOTAL_FILL  = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    thin = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")
    file = f"hisobot_{month}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"
    conn_local = sqlite3.connect("ishchilar.db")
    cursor_local = conn_local.cursor()
    cursor_local.execute("SELECT fullname, passport FROM ishchilar WHERE role='worker'")
    workers = cursor_local.fetchall()
    start_date = datetime.strptime(month + "-01", "%Y-%m-%d").date()
    end_date = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    days_list = []
    cur_d = start_date
    while cur_d <= end_date:
        days_list.append(cur_d)
        cur_d += timedelta(days=1)
    total_cols = 2 + len(days_list) + 2
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    t = ws["A1"]
    t.value = f"DAVOMAT JADVALI — {month}"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = HEADER_FILL
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28
    headers = ["Ism", "Familiya"]
    for d in days_list:
        lbl = str(d.day).zfill(2)
        if not is_work_day(d):
            lbl += "*"
        headers.append(lbl)
    headers += ["Kelgan", "Kelmagan"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 20
    ish_kunlari = count_work_days_in_month(month)
    for ri, (fullname, passport) in enumerate(workers, 3):
        parts = fullname.split(" ")
        ism = parts[0]
        familya = parts[1] if len(parts) > 1 else ""
        cursor_local.execute(
            "SELECT date FROM davomat WHERE passport=? AND date LIKE ? AND confirmed=1",
            (passport, f"{month}%")
        )
        kelgan_sanalar = {r[0] for r in cursor_local.fetchall()}
        for ci, col_name in enumerate(["ism", "familya"], 1):
            val = ism if ci == 1 else familya
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", bold=True, size=10)
            c.fill = NAME_FILL
            c.alignment = LEFT
            c.border = BORDER
        for ci, d in enumerate(days_list, 3):
            cell = ws.cell(row=ri, column=ci)
            d_str = d.strftime("%Y-%m-%d")
            if not is_work_day(d):
                cell.value = ""
                cell.fill = RED_FILL
            elif d_str in kelgan_sanalar:
                cell.value = "✓"
                cell.fill = GREEN_FILL
                cell.font = Font(name="Arial", bold=True, color="1A6B1A", size=11)
            else:
                cell.value = "✗"
                cell.fill = YELLOW_FILL
                cell.font = Font(name="Arial", bold=True, color="8B4513", size=11)
            cell.alignment = CENTER
            cell.border = BORDER
        kelgan_ish = len([d for d in days_list if d.strftime("%Y-%m-%d") in kelgan_sanalar and is_work_day(d)])
        kelmagan_cnt = max(ish_kunlari - kelgan_ish, 0)
        kc = ws.cell(row=ri, column=2 + len(days_list) + 1, value=kelgan_ish)
        kc.fill = TOTAL_FILL
        kc.font = Font(name="Arial", bold=True, color="227B22", size=11)
        kc.alignment = CENTER
        kc.border = BORDER
        kc2 = ws.cell(row=ri, column=2 + len(days_list) + 2, value=kelmagan_cnt)
        kc2.fill = TOTAL_FILL
        kc2.font = Font(name="Arial", bold=True, color="CC0000", size=11)
        kc2.alignment = CENTER
        kc2.border = BORDER
        ws.row_dimensions[ri].height = 18
    conn_local.close()
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    for i in range(len(days_list)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 4.5
    ws.column_dimensions[get_column_letter(3 + len(days_list))].width = 9
    ws.column_dimensions[get_column_letter(4 + len(days_list))].width = 11
    lr = len(workers) + 4
    ws.merge_cells(f"A{lr}:F{lr}")
    lc = ws.cell(row=lr, column=1, value="Yashil (✓) = Kelgan   |   Sariq (✗) = Kelmagan (ish kuni)   |   Qizil = Dam olish / Bayram")
    lc.font = Font(name="Arial", italic=True, size=8, color="555555")
    ws.freeze_panes = "C3"
    wb.save(file)
    return file

def pdf_month_report(month=None):
    days, workers = get_month_days_and_attendance(month)
    month_str = month if month else datetime.now().strftime("%Y-%m")
    file_name = f"hisobot_{month_str}.pdf"
    c = canvas.Canvas(file_name, pagesize=A4)
    y = 800
    c.drawString(50, y, f"Oylik Davomat Hisoboti: {month_str}")
    y -= 20
    ish_kunlari = count_work_days_in_month(month_str)
    c.drawString(50, y, f"Jami ish kunlari: {ish_kunlari}")
    y -= 30
    for fullname, _, _, kelgan_sanalar in workers:
        kelgan_ish = len([d for d in days if d in kelgan_sanalar and is_work_day(datetime.strptime(d, "%Y-%m-%d").date())])
        kelmagan = max(ish_kunlari - kelgan_ish, 0)
        line = f"{fullname}: Kelgan={kelgan_ish}, Kelmagan={kelmagan}"
        c.drawString(50, y, line)
        y -= 20
        if y < 50:
            c.showPage()
            y = 800
    c.save()
    return file_name

def export_oylik_maosh_excel(month=None):
    month_str = month if month else datetime.now().strftime("%Y-%m")
    days, workers = get_month_days_and_attendance(month_str)
    ish_kunlari = count_work_days_in_month(month_str)
    wb = Workbook()
    ws = wb.active
    ws.title = "Umumiy Oylik Maosh"
    
    HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    DATA_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    
    headers = ["F.I.Sh", "Oylik (nominal)", "Ish Kunlari", "Kelgan", "Kelmagan", "Kechikish (Daq.)", "Jarima", "Avans", "Sof Maosh"]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions["A"].width = 30

    conn_local = sqlite3.connect("ishchilar.db")
    cur = conn_local.cursor()
    row_idx = 2
    for fullname, passport, salary, kelgan_sanalar in workers:
        cur.execute("SELECT SUM(late_minutes), SUM(jarima) FROM moliya WHERE passport=? AND date LIKE ?", (passport, f"{month_str}%"))
        moliya_data = cur.fetchone()
        total_late = moliya_data[0] or 0
        total_jarima = moliya_data[1] or 0
        cur.execute("SELECT SUM(avans) FROM moliya WHERE passport=? AND date LIKE ?", (passport, f"{month_str}%"))
        total_avans = cur.fetchone()[0] or 0
        kelgan_ish = len([d for d in days if d in kelgan_sanalar and is_work_day(datetime.strptime(d, "%Y-%m-%d").date())])
        kelmagan = max(ish_kunlari - kelgan_ish, 0)
        bir_kunlik = salary / ish_kunlari if ish_kunlari else 0
        yoqchilik = kelmagan * bir_kunlik
        
        sof = salary - yoqchilik - total_avans
        
        row_data = [fullname, salary, ish_kunlari, kelgan_ish, kelmagan, total_late, int(total_jarima), int(total_avans), int(sof)]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = CENTER
            if row_idx % 2 == 0:
                cell.fill = DATA_FILL
            if col_idx == 9: # Sof maosh
                cell.font = Font(bold=True, color="006100")
                cell.fill = PatternFill("solid", start_color="C6E0B4", end_color="C6E0B4")
        row_idx += 1
        
    conn_local.close()
    file_name = f"oylik_maosh_{month_str}.xlsx"
    wb.save(file_name)
    return file_name

# ================== 6. START / SELECTION ==================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = ReplyKeyboardMarkup([["👔 Boshliq"], ["👷 Ishchi"]], resize_keyboard=True)
    await update.message.reply_text("Kim sifatida kirasiz?", reply_markup=kb)

async def role_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    if "Boshliq" in text:
        if update.effective_user.id != BOSS_ID:
            await update.message.reply_text("⛔ Siz boshliq emassiz")
            return ConversationHandler.END
        context.user_data["role"] = "boshliq"
        await boshliq_menu(update, context)
    else:
        context.user_data["role"] = "ishchi"
        await ishchi_main_menu(update, context)
    return ConversationHandler.END

# ================== 7. MENULAR ==================
async def boshliq_menu(update, context):
    kb = ReplyKeyboardMarkup(
        [
            ["📍 Boss lokatsiya", "💰 Oylik maosh"],
            ["📄 PDF chiqarish", "📊 Excel chiqarish"],
            ["📆 Kunlik", "📆 Oylik"],
            ["➕ Ishchi qo'shish", "❌ Ishchi o'chirish"],
            ["📨 Ish berish", "📋 Ma'lumot"],
            ["📌 Eslatmalar", "⬅️ Orqaga"]
        ],
        resize_keyboard=True
    )
    msg = "👔 Boshliq paneli"
    if update.callback_query:
        await update.callback_query.message.reply_text(msg, reply_markup=kb)
    else:
        await update.message.reply_text(msg, reply_markup=kb)

async def ishchi_main_menu(update, context):
    kb = ReplyKeyboardMarkup(
        [
            ["✅ Keldim", "💰 Oylik maosh"],
            ["📋 Davomat", "📩 Topshiriqlar"],
            ["⬅️ Orqaga"]
        ],
        resize_keyboard=True
    )
    msg = "👷 Ishchi paneli"
    if update.callback_query:
        await update.callback_query.message.reply_text(msg, reply_markup=kb)
    else:
        await update.message.reply_text(msg, reply_markup=kb)

# ================== 8. MA'LUMOT (Boshliq) ==================
async def malumot_handler(update, context):
    if update.effective_user.id != BOSS_ID:
        await update.message.reply_text("⛔ Ruxsat yo'q")
        return
    cursor.execute("SELECT fullname, passport, jshshir, birthdate, daily_salary, pin, code, location FROM ishchilar WHERE role='worker'")
    workers = cursor.fetchall()
    if not workers:
        await update.message.reply_text("❌ Ishchilar mavjud emas.")
        return
    month_str = datetime.now().strftime("%Y-%m")
    ish_kunlari = count_work_days_in_month(month_str)
    text = f"📋 <b>Ishchilar ma'lumotlari</b>\n📅 Bu oy ish kunlari: {ish_kunlari}\n\n"
    for i, (fullname, passport, jshshir, birthdate, salary, pin, code, location) in enumerate(workers, 1):
        bir_kunlik = salary // ish_kunlari if ish_kunlari else 0
        bir_soatlik = bir_kunlik // WORK_HOURS_PER_DAY
        text += (
            f"━━━━━━━━━━━━━━━━\n"
            f"<b>{i}. {fullname}</b>\n"
            f"🪪 Pasport: <code>{passport}</code>\n"
            f"🆔 JSHSHIR: <code>{jshshir or '—'}</code>\n"
            f"🎂 Tug'ilgan sana: {birthdate or '—'}\n"
            f"💰 Oylik: <b>{salary:,}</b> so'm\n"
            f"📆 Bir kunlik: {bir_kunlik:,} so'm\n"
            f"⏱ Bir soatlik: {bir_soatlik:,} so'm\n"
            f"🔐 Kirish kodi: <code>{code}</code>\n"
            f"🔒 PIN: <code>{pin}</code>\n"
            f"📍 Lokatsiya: {location or '—'}\n"
        )
    await update.message.reply_text(text, parse_mode="HTML")

# ================== 9. BOSS LOKATSIYA ==================
async def ask_boss_location(update, context):
    await update.message.reply_text(
        "📍 Boss lokatsiyasini kiriting (format: lat,long)\nMasalan: 40.861751,71.458907"
    )
    return BOSS_LOCATION

async def save_boss_location(update, context):
    text = update.message.text.strip()
    menu_tugmalari = [
        "📍 Boss lokatsiya", "💰 Oylik maosh", "📄 PDF chiqarish",
        "📊 Excel chiqarish", "📆 Kunlik", "📆 Oylik",
        "➕ Ishchi qo'shish", "❌ Ishchi o'chirish", "📨 Ish berish",
        "✅ Keldim", "📋 Davomat", "⬅️ Orqaga", "📋 Ma'lumot", "📩 Topshiriqlar", "📌 Eslatmalar"
    ]
    if text in menu_tugmalari:
        await update.message.reply_text("⚠️ Lokatsiya kiritish bekor qilindi.")
        await boshliq_menu(update, context)
        return ConversationHandler.END
    if "," not in text:
        await update.message.reply_text("❌ Format xato! lat,long formatda kiriting:")
        return BOSS_LOCATION
    try:
        lat_str, lon_str = text.split(",")
        lat = float(lat_str.strip())
        lon = float(lon_str.strip())
    except ValueError:
        await update.message.reply_text("❌ Xato koordinatalar. Qayta kiriting:")
        return BOSS_LOCATION
    cursor.execute("""
        INSERT INTO boss_location (id, map_link) VALUES (1, ?)
        ON CONFLICT(id) DO UPDATE SET map_link=excluded.map_link
    """, (f"{lat},{lon}",))
    conn.commit()
    await update.message.reply_text(f"✅ Boss lokatsiyasi yangilandi: {lat},{lon}")
    await boshliq_menu(update, context)
    return ConversationHandler.END

# ================== ESLATMALAR QISMI (YANGI) ==================
def get_eslatmalar_markup():
    conn_local = sqlite3.connect("ishchilar.db")
    cursor_local = conn_local.cursor()
    # Bajarilmaganlar ustda (is_done=0), bajarilganlar pastda (is_done=1)
    cursor_local.execute("SELECT id, text, is_done FROM eslatmalar WHERE boss_id=? ORDER BY is_done ASC, id DESC", (BOSS_ID,))
    eslatmalar = cursor_local.fetchall()
    conn_local.close()

    keyboard = []
    for es_id, text, is_done in eslatmalar:
        short_text = (text[:30] + '..') if len(text) > 30 else text
        if is_done:
            # ✅ holatidagi tugma. Boss yana bossa O'CHIRIB YUBORILADI
            btn_text = f"✅ {short_text} ❌"
            keyboard.append([InlineKeyboardButton(btn_text, callback_data=f"del_eslatma_{es_id}")])
        else:
            # ◻️ holatidagi tugma. Boss bossa bajarildi belgilanib pastga tushadi
            btn_text = f"◻️ {short_text}"
            keyboard.append([InlineKeyboardButton(btn_text, callback_data=f"toggle_eslatma_{es_id}")])

    keyboard.append([InlineKeyboardButton("➕ Yangi eslatma qo'shish", callback_data="add_eslatma")])
    return InlineKeyboardMarkup(keyboard)

async def eslatmalar_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != BOSS_ID:
        return
    markup = get_eslatmalar_markup()
    msg = "📋 <b>Sizning eslatmalaringiz:</b>\n<i>- Ishni bajarilgach ustiga bosing (✅ bo'ladi va pastga tushadi)\n- ✅ ni yana bir marta bossangiz butunlay o'chib ketadi.</i>"
    if update.callback_query:
        await update.callback_query.message.edit_text(msg, parse_mode="HTML", reply_markup=markup)
    else:
        await update.message.reply_text(msg, parse_mode="HTML", reply_markup=markup)

async def handle_eslatma_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "add_eslatma":
        await query.message.reply_text("✍️ Eslatma matnini kiriting (Masalan: Ertaga 10:00 da majlis):", reply_markup=ReplyKeyboardMarkup([["⬅️ Bekor qilish"]], resize_keyboard=True))
        return YANGI_ESLATMA
        
    elif data.startswith("toggle_eslatma_"):
        es_id = int(data.split("_")[2])
        conn_local = sqlite3.connect("ishchilar.db")
        cursor_local = conn_local.cursor()
        # Ishni bajarildi (1) qilib qo'yish
        cursor_local.execute("UPDATE eslatmalar SET is_done=1 WHERE id=?", (es_id,))
        conn_local.commit()
        conn_local.close()
        
        markup = get_eslatmalar_markup()
        await query.message.edit_reply_markup(reply_markup=markup)

    elif data.startswith("del_eslatma_"):
        es_id = int(data.split("_")[2])
        conn_local = sqlite3.connect("ishchilar.db")
        cursor_local = conn_local.cursor()
        # Bajarilgan ish yana bosilsa, uni umuman ro'yxatdan o'chirish
        cursor_local.execute("DELETE FROM eslatmalar WHERE id=?", (es_id,))
        conn_local.commit()
        conn_local.close()
        
        markup = get_eslatmalar_markup()
        await query.message.edit_reply_markup(reply_markup=markup)

async def save_yangi_eslatma(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "⬅️ Bekor qilish":
        await boshliq_menu(update, context)
        return ConversationHandler.END

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn_local = sqlite3.connect("ishchilar.db")
    cursor_local = conn_local.cursor()
    cursor_local.execute("INSERT INTO eslatmalar (boss_id, text, is_done, created_at) VALUES (?, ?, 0, ?)", (BOSS_ID, text, now_str))
    conn_local.commit()
    conn_local.close()

    await update.message.reply_text("✅ Eslatma saqlandi!")
    await boshliq_menu(update, context)
    
    markup = get_eslatmalar_markup()
    await update.message.reply_text("📋 <b>Sizning eslatmalaringiz:</b>", parse_mode="HTML", reply_markup=markup)
    return ConversationHandler.END

# ================== 10. ISHCHI KELDIM ==================
async def keldim_code_start(update, context):
    today = datetime.now().date()
    if not is_work_day(today):
        day_names = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
        day_name = day_names[today.weekday()]
        if today.weekday() >= 5:
            await update.message.reply_text(f"🔵 Bugun {day_name} — dam olish kuni. Ish kuni emas!")
        else:
            await update.message.reply_text(f"🔵 Bugun davlat bayrami — dam olish kuni.")
        await ishchi_main_menu(update, context)
        return ConversationHandler.END
    kb = ReplyKeyboardMarkup(
        [[KeyboardButton("📍 Joyim", request_location=True)], ["⬅️ Orqaga"]],
        resize_keyboard=True
    )
    await update.message.reply_text("🟢 Joylashuvni (📍 Joyim tugmasi orqali) yuboring:", reply_markup=kb)
    return KELDIM_CODE

async def check_code_or_location(update, context):
    if update.message.location:
        context.user_data["location_now"] = (
            update.message.location.latitude,
            update.message.location.longitude
        )
        await update.message.reply_text("📍 Lokatsiya qabul qilindi. Endi 6 xonali ishchi kodingizni kiriting:")
        return KELDIM_CODE
    code_text = update.message.text.strip()
    if code_text == "⬅️ Orqaga":
        await ishchi_main_menu(update, context)
        return ConversationHandler.END
    try:
        code = int(code_text)
    except ValueError:
        await update.message.reply_text("❌ Iltimos faqat raqamli kodni kiriting:")
        return KELDIM_CODE
    cursor.execute("SELECT passport, fullname, location FROM ishchilar WHERE code=?", (code,))
    result = cursor.fetchone()
    if not result:
        await update.message.reply_text("❌ Kod topilmadi, qayta kiriting:")
        return KELDIM_CODE
    passport, fullname, assigned_location = result
    context.user_data.update({
        "passport": passport,
        "fullname": fullname,
        "assigned_location": assigned_location
    })
    await update.message.reply_text(f"👋 {fullname}, xush kelibsiz!\n🔒 Iltimos, PIN kodni kiriting:")
    return PIN_STEP

async def check_pin_for_keldim(update, context):
    pin = update.message.text.strip()
    passport = context.user_data.get("passport")
    cursor.execute("SELECT fullname, location, daily_salary FROM ishchilar WHERE passport=? AND pin=?", (passport, pin))
    row = cursor.fetchone()
    if not row:
        await update.message.reply_text("❌ PIN noto'g'ri, qayta kiriting:")
        return PIN_STEP
    fullname, assigned_location, daily_salary = row
    user_location = context.user_data.get("location_now")
    if not user_location:
        await update.message.reply_text("❌ Avval joylashuvingizni yuborishingiz kerak.")
        return ConversationHandler.END
    user_lat, user_lon = user_location
    cursor.execute("SELECT map_link FROM boss_location WHERE id=1")
    boss_row = cursor.fetchone()
    if not boss_row or not boss_row[0]:
        await update.message.reply_text("❌ Boss lokatsiyasi o'rnatilmagan.")
        return ConversationHandler.END
    boss_lat, boss_lon = map(float, boss_row[0].split(","))
    if isinstance(assigned_location, str) and assigned_location.startswith("http"):
        assigned_lat, assigned_lon = parse_latlon_from_link(assigned_location)
    else:
        try:
            assigned_lat, assigned_lon = map(float, assigned_location.split(","))
        except:
            assigned_lat, assigned_lon = None, None
    if assigned_lat is None:
        await update.message.reply_text("❌ Ish joyingiz koordinatalari noto'g'ri sozlangan.")
        return ConversationHandler.END
    distance_to_work = calculate_distance(user_lat, user_lon, assigned_lat, assigned_lon) * 1000
    distance_to_boss = calculate_distance(user_lat, user_lon, boss_lat, boss_lon) * 1000
    if distance_to_work > 100:
        await update.message.reply_text(f"❌ Siz ish joyidan uzoqdasiz.\n📏 Masofa: {int(distance_to_work)} metr (Ruxsat: 100m ichida)")
        return ConversationHandler.END
    if distance_to_boss > 100:
        await update.message.reply_text(f"❌ Siz boss turgan joydan uzoqdasiz.\n📏 Masofa: {int(distance_to_boss)} metr")
        return ConversationHandler.END
    today = datetime.now().strftime("%Y-%m-%d")
    now_time = datetime.now().strftime("%H:%M")
    cursor.execute("SELECT id FROM davomat WHERE passport=? AND date=?", (passport, today))
    if cursor.fetchone():
        await update.message.reply_text("⚠️ Siz bugun allaqachon davomatdan o'tgansiz.")
        await ishchi_main_menu(update, context)
        return ConversationHandler.END
    cursor.execute("INSERT INTO davomat (passport, date, time, confirmed) VALUES (?, ?, ?, 1)", (passport, today, now_time))
    conn.commit()
    month_str = datetime.now().strftime("%Y-%m")
    fine_info = calculate_late_fine(passport, now_time, daily_salary, month_str)
    late_minutes = fine_info["late_minutes"]
    jarima = fine_info["fine"]
    if late_minutes > 0:
        cursor.execute("""
            INSERT INTO moliya (passport, date, late_minutes, jarima) VALUES (?, ?, ?, ?)
            ON CONFLICT DO NOTHING
        """, (passport, today, late_minutes, jarima))
        conn.commit()
    late_str = f"{late_minutes // 60} soat {late_minutes % 60} daqiqa" if late_minutes > 0 else "Yo'q"
    jarima_str = f"{jarima:,} so'm" if jarima > 0 else "Yo'q"
    msg_ishchi = (
        f"✅ Keldim qabul qilindi!\n"
        f"📅 Sana: {today}\n"
        f"⏰ Kelish vaqti: {now_time}\n"
    )
    if late_minutes > 0:
        msg_ishchi += f"⏱ Kechikish: {late_str}\n💸 Jarima: {jarima_str}"
    else:
        msg_ishchi += "✅ Vaqtida keldingiz!"
    await update.message.reply_text(msg_ishchi)
    try:
        await context.bot.send_message(
            chat_id=BOSS_ID,
            text=(
                f"📢 <b>Keldim xabarnomasi</b>\n\n"
                f"👤 Ishchi: {fullname}\n"
                f"📅 Sana: {today}\n"
                f"⏰ Vaqt: {now_time}\n"
                f"⏱ Kechikish: {late_str}\n"
                f"💸 Jarima: {jarima_str}"
            ),
            parse_mode="HTML"
        )
    except Exception as e:
        print(f"Bossga xabar yuborishda xatolik: {e}")
    await ishchi_main_menu(update, context)
    return ConversationHandler.END

# ================== 11. ISH BERISH ==================
async def ish_berish_start(update, context):
    if update.effective_user.id != BOSS_ID:
        await update.message.reply_text("⛔ Sizda ruxsat yo'q")
        return ConversationHandler.END
    cursor.execute("SELECT id, fullname, passport FROM ishchilar WHERE role='worker'")
    workers = cursor.fetchall()
    if not workers:
        await update.message.reply_text("❌ Ishchilar topilmadi.")
        return ConversationHandler.END

    buttons = []
    for wid, name, passport in workers:
        cursor.execute("""
            SELECT COUNT(*) FROM messages
            WHERE ishchi_id=? AND sender='ishchi' AND is_read=0
        """, (wid,))
        unread = cursor.fetchone()[0] or 0
        label = f"👤 {name}"
        if unread > 0:
            label += f" 🔴{unread}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"ish_{wid}")])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="back_boss")])
    await update.message.reply_text(
        "📋 Topshiriq berish uchun ishchi tanlang:\n(🔴 = o'qilmagan javoblar)",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
    return ISHBERISH_SELECT

async def ish_berish_tanlash(update, context):
    q = update.callback_query
    await q.answer()
    if q.data == "back_boss":
        await boshliq_menu(q, context)
        return ConversationHandler.END

    ishchi_id = int(q.data.split("_")[1])
    context.user_data["ishchi_id"] = ishchi_id

    cursor.execute("SELECT fullname, tg_id, passport FROM ishchilar WHERE id=?", (ishchi_id,))
    row = cursor.fetchone()
    if not row:
        await q.message.reply_text("❌ Ishchi topilmadi.")
        return ConversationHandler.END
    fullname, tg_id, passport = row
    context.user_data.update({
        "ishchi_tg_id": tg_id,
        "ishchi_passport": passport,
        "ish_berish_buffer": [],
        "ish_berish_name": fullname
    })

    cursor.execute("""
        UPDATE messages SET is_read=1
        WHERE ishchi_id=? AND sender='ishchi' AND is_read=0
    """, (ishchi_id,))
    conn.commit()

    cursor.execute("""
        SELECT sender, type, content, date FROM messages
        WHERE ishchi_id=? ORDER BY id DESC LIMIT 10
    """, (ishchi_id,))
    history = cursor.fetchall()[::-1]

    text = f"📬 <b>{fullname}</b> bilan topshiriqlar tarixi:\n\n"
    if history:
        for sender, mtype, content, mdate in history:
            who = "👔 Siz" if sender == "boss" else f"👷 {fullname}"
            if mtype == "voice":
                text += f"[{mdate}] {who}: 🎤 Ovozli xabar\n"
            else:
                short = content[:80] + ("..." if len(content) > 80 else "")
                text += f"[{mdate}] {who}: {short}\n"
        text += "\n"
    else:
        text += "Hali topshiriq yuborilmagan.\n\n"

    text += "✍️ Yangi topshiriq yozing yoki 🎤 ovozli xabar yuboring.\nTugatgach '📤 Yuborish' tugmasini bosing."
    await q.message.reply_text(
        text,
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            [["📤 Yuborish"], ["⬅️ Bekor qilish"]], resize_keyboard=True
        )
    )
    return ISHBERISH_MESSAGE

async def ish_berish_yuborish(update, context):
    if update.message.text == "⬅️ Bekor qilish":
        context.user_data["ish_berish_buffer"] = []
        await boshliq_menu(update, context)
        return ConversationHandler.END

    buffer = context.user_data.get("ish_berish_buffer", [])

    if update.message.text == "📤 Yuborish":
        tg_id = context.user_data.get("ishchi_tg_id")
        ishchi_id = context.user_data.get("ishchi_id")
        passport = context.user_data.get("ishchi_passport")
        if not buffer:
            await update.message.reply_text("❌ Hech qanday xabar yozilmadi.")
            return ISHBERISH_MESSAGE
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        for msg in buffer:
            try:
                if msg["type"] == "voice":
                    sent = await context.bot.send_voice(
                        chat_id=tg_id, voice=msg["content"],
                        caption="📢 Boshliqdan topshiriq\n\n💬 Javob berish uchun /topshiriqlar buyrug'ini yoki «📩 Topshiriqlar» tugmasini bosing"
                    )
                else:
                    sent = await context.bot.send_message(
                        chat_id=tg_id,
                        text=f"📢 <b>Boshliqdan yangi topshiriq:</b>\n\n{msg['content']}\n\n💬 Javob berish uchun «📩 Topshiriqlar» tugmasini bosing",
                        parse_mode="HTML"
                    )
            except Exception as e:
                print(f"Ishchiga xabar borishda xato: {e}")
            cursor.execute("""
                INSERT INTO messages (ishchi_id, passport, sender, type, content, date, is_read)
                VALUES (?, ?, 'boss', ?, ?, ?, 0)
            """, (ishchi_id, passport, msg["type"], msg["content"], now_str))
        conn.commit()
        context.user_data["ish_berish_buffer"] = []
        await update.message.reply_text("✅ Topshiriqlar muvaffaqiyatli yetkazildi.")
        await boshliq_menu(update, context)
        return ConversationHandler.END

    if update.message.voice:
        buffer.append({"type": "voice", "content": update.message.voice.file_id})
        await update.message.reply_text("🎤 Ovozli topshiriq kiritildi. Yana qo'shishingiz yoki 📤 Yuborish bosishingiz mumkin.")
    elif update.message.text:
        buffer.append({"type": "text", "content": update.message.text})
        await update.message.reply_text("✍️ Matnli topshiriq kiritildi. Yana qo'shishingiz yoki 📤 Yuborish bosishingiz mumkin.")

    context.user_data["ish_berish_buffer"] = buffer
    return ISHBERISH_MESSAGE

# ================== 12. TOPSHIRIQLAR (Ishchi) ==================
async def ishchi_topshiriqlar(update, context):
    passport = context.user_data.get("passport")
    if not passport:
        await update.message.reply_text("❌ Avval ✅ Keldim orqali tizimga kiring.")
        return ConversationHandler.END

    cursor.execute("SELECT id FROM ishchilar WHERE passport=?", (passport,))
    row = cursor.fetchone()
    if not row:
        await update.message.reply_text("❌ Ishchi topilmadi.")
        return ConversationHandler.END
    ishchi_id = row[0]

    cursor.execute("""
        SELECT id, sender, type, content, date, is_read FROM messages
        WHERE ishchi_id=? ORDER BY id DESC LIMIT 20
    """, (ishchi_id,))
    msgs = cursor.fetchall()[::-1]

    if not msgs:
        await update.message.reply_text(
            "📭 Hali topshiriq yo'q.",
            reply_markup=ReplyKeyboardMarkup(
                [["⬅️ Orqaga"]], resize_keyboard=True
            )
        )
        return ConversationHandler.END

    cursor.execute("""
        UPDATE messages SET is_read=1
        WHERE ishchi_id=? AND sender='boss' AND is_read=0
    """, (ishchi_id,))
    conn.commit()

    text = "📩 <b>Topshiriqlar tarixi:</b>\n\n"
    for msg_id, sender, mtype, content, mdate, is_read in msgs:
        who = "👔 Boshliq" if sender == "boss" else "👷 Siz"
        unread_mark = " 🔴" if sender == "boss" and not is_read else ""
        if mtype == "voice":
            text += f"[{mdate}] {who}{unread_mark}: 🎤 Ovozli xabar\n"
        else:
            short = content[:100] + ("..." if len(content) > 100 else "")
            text += f"[{mdate}] {who}{unread_mark}: {short}\n"

    text += "\n💬 Boshliqqà javob yozmoqchi bo'lsangiz, xabarni yuboring:"

    context.user_data["topshiriq_reply_mode"] = True
    context.user_data["topshiriq_ishchi_id"] = ishchi_id

    await update.message.reply_text(
        text,
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            [["⬅️ Orqaga"]], resize_keyboard=True
        )
    )
    return TOPSHIRIQ_JAVOB

async def ishchi_topshiriq_javob(update, context):
    text = update.message.text

    if text == "⬅️ Orqaga":
        context.user_data.pop("topshiriq_reply_mode", None)
        await ishchi_main_menu(update, context)
        return ConversationHandler.END

    passport = context.user_data.get("passport")
    ishchi_id = context.user_data.get("topshiriq_ishchi_id")

    if not passport or not ishchi_id:
        await ishchi_main_menu(update, context)
        return ConversationHandler.END

    cursor.execute("SELECT fullname FROM ishchilar WHERE passport=?", (passport,))
    row = cursor.fetchone()
    fullname = row[0] if row else "Noma'lum"

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    if update.message.voice:
        cursor.execute("""
            INSERT INTO messages (ishchi_id, passport, sender, type, content, date, is_read)
            VALUES (?, ?, 'ishchi', 'voice', ?, ?, 0)
        """, (ishchi_id, passport, update.message.voice.file_id, now_str))
        conn.commit()
        try:
            await context.bot.send_voice(
                chat_id=BOSS_ID,
                voice=update.message.voice.file_id,
                caption=f"📩 <b>{fullname}</b> dan ovozli javob\n🕐 {now_str}",
                parse_mode="HTML"
            )
        except Exception as e:
            print(f"Bossga voice xabar yuborishda xato: {e}")
        await update.message.reply_text("✅ Ovozli javobingiz boshliqqa yuborildi.")
    elif update.message.text and text != "⬅️ Orqaga":
        cursor.execute("""
            INSERT INTO messages (ishchi_id, passport, sender, type, content, date, is_read)
            VALUES (?, ?, 'ishchi', 'text', ?, ?, 0)
        """, (ishchi_id, passport, text, now_str))
        conn.commit()
        try:
            await context.bot.send_message(
                chat_id=BOSS_ID,
                text=f"📩 <b>{fullname}</b> dan javob:\n\n{text}\n\n🕐 {now_str}",
                parse_mode="HTML"
            )
        except Exception as e:
            print(f"Bossga xabar yuborishda xato: {e}")
        await update.message.reply_text("✅ Javobingiz boshliqqa yuborildi.")

    return TOPSHIRIQ_JAVOB

# ================== 13. ISHCHI QO'SHISH ==================
async def ishchi_qoshish(update, context):
    await update.message.reply_text("👤 Yangi ishchining Ism va Familiyasini kiriting:")
    return FULLNAME

async def ishchi_fullname(update, context):
    context.user_data["fullname"] = update.message.text.title()
    await update.message.reply_text("🪪 Pasport seriya va raqamini kiriting (masalan: AA1234567):")
    return PASSPORT

async def ishchi_passport(update, context):
    passport = update.message.text.upper().replace(" ", "")
    if not re.match(r"^[A-Z]{2}\d{7}$", passport):
        await update.message.reply_text("❌ Pasport formati noto'g'ri!\nIltimos, 2 ta harf va 7 ta raqam kiriting (Masalan: AA1234567):")
        return PASSPORT
        
    cursor.execute("SELECT id FROM ishchilar WHERE passport=?", (passport,))
    if cursor.fetchone():
        await update.message.reply_text("⚠️ Bu pasport bazada bor. Boshqa pasport kiriting:")
        return PASSPORT
        
    context.user_data["passport"] = passport
    await update.message.reply_text("🆔 Endi ishchining 14 xonali JSHSHIR raqamini kiriting:")
    return JSHSHIR

async def ishchi_jshshir(update, context):
    jshshir = update.message.text.strip()
    if len(jshshir) != 14 or not jshshir.isdigit():
        await update.message.reply_text("❌ JSHSHIR aynan 14 ta raqamdan iborat bo'lishi kerak. Qayta kiriting:")
        return JSHSHIR
        
    idx = jshshir[0]
    dd = jshshir[1:3]
    mm = jshshir[3:5]
    yy = jshshir[5:7]
    
    if idx in ['1', '2']: year = "18" + yy
    elif idx in ['3', '4']: year = "19" + yy
    elif idx in ['5', '6']: year = "20" + yy
    else: year = "20" + yy
        
    birthdate = f"{dd}.{mm}.{year}"
    context.user_data["jshshir"] = jshshir
    context.user_data["birthdate"] = birthdate
    
    await update.message.reply_text(f"✅ Tug'ilgan sana avtomatik aniqlandi: {birthdate}\n\n💰 Endi bir oylik nominal maoshini kiriting (faqat raqam, so'mda):")
    return MAOSH

async def ishchi_maosh(update, context):
    try:
        salary = int(update.message.text.replace(" ", "").replace(",", ""))
    except ValueError:
        await update.message.reply_text("❌ Xato. Iltimos faqat raqam kiriting:")
        return MAOSH
    context.user_data["salary"] = salary
    await update.message.reply_text("🔒 Ishchi uchun 4 xonali PIN kod belgilang (masalan: 1234):")
    return PIN_ADD

async def ishchi_pin_add(update, context):
    pin = update.message.text.strip()
    if not pin.isdigit() or len(pin) != 4:
        await update.message.reply_text("❌ PIN 4 ta raqamdan iborat bo'lishi kerak. Qayta kiriting:")
        return PIN_ADD
    context.user_data["pin"] = pin
    
    await update.message.reply_text(
        "📍 Ishchi doimiy ishlash joyi koordinatasini kiriting (lat,long formatda):\n"
        "Masalan: 40.8625,71.4575"
    )
    return LOCATION

async def ishchi_location(update, context):
    location = update.message.text.strip()
    fullname = context.user_data["fullname"]
    passport = context.user_data["passport"]
    jshshir = context.user_data.get("jshshir", "")
    birthdate = context.user_data.get("birthdate", "")
    salary = context.user_data["salary"]
    pin = context.user_data.get("pin", "0000")
    code = generate_code()
    
    cursor.execute("""
        INSERT INTO ishchilar (fullname, passport, jshshir, birthdate, code, role, daily_salary, pin, location)
        VALUES (?, ?, ?, ?, ?, 'worker', ?, ?, ?)
    """, (fullname, passport, jshshir, birthdate, code, salary, pin, location))
    conn.commit()
    
    month_str = datetime.now().strftime("%Y-%m")
    ish_kunlari = count_work_days_in_month(month_str)
    bir_kunlik = salary / ish_kunlari if ish_kunlari else 0
    bir_soatlik = bir_kunlik / WORK_HOURS_PER_DAY
    
    await update.message.reply_text(
        f"✅ Ishchi muvaffaqiyatli qo'shildi!\n\n"
        f"👤 F.I.Sh: {fullname}\n"
        f"🪪 Pasport: {passport}\n"
        f"🆔 JSHSHIR: {jshshir}\n"
        f"🎂 Tug'ilgan sana: {birthdate}\n"
        f"💰 Oylik: {salary:,} so'm\n"
        f"📅 Bu oyda ish kunlari: {ish_kunlari} kun\n"
        f"📆 Bir kunlik: {int(bir_kunlik):,} so'm\n"
        f"⏱ Bir soatlik: {int(bir_soatlik):,} so'm\n"
        f"📍 Lokatsiya: {location}\n"
        f"🔒 PIN: {pin}\n"
        f"🔐 <b>Kirish kodi: {code}</b>",
        parse_mode="HTML"
    )
    await boshliq_menu(update, context)
    return ConversationHandler.END

# ================== 14. ISHCHI O'CHIRISH ==================
async def ishchi_ochirish(update, context):
    conn_local = sqlite3.connect("ishchilar.db")
    cursor_local = conn_local.cursor()
    
    cursor_local.execute("SELECT id, fullname FROM ishchilar WHERE role='worker' ORDER BY id")
    workers = cursor_local.fetchall()
    conn_local.close()

    if not workers:
        if update.callback_query:
            await update.callback_query.message.reply_text("Ishchilar mavjud emas ❌")
        else:
            await update.message.reply_text("Ishchilar mavjud emas ❌")
        return

    buttons = [[InlineKeyboardButton(f"🗑 {n}", callback_data=f"del_{i}")] for i, n in workers]
    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="back_boss")])

    markup = InlineKeyboardMarkup(buttons)
    if update.callback_query:
        try:
            await update.callback_query.message.edit_reply_markup(reply_markup=markup)
        except Exception:
            await update.callback_query.message.reply_text("O'chirish uchun ishchini tanlang:", reply_markup=markup)
    else:
        await update.message.reply_text("O'chirish uchun ishchini tanlang:", reply_markup=markup)

async def delete_worker(update, context):
    query = update.callback_query
    await query.answer()

    if query.data == "back_boss":
        await boshliq_menu(update, context)
        return

    try:
        worker_id = int(query.data.split("_")[1])
    except (IndexError, ValueError):
        return

    conn_local = sqlite3.connect("ishchilar.db")
    cursor_local = conn_local.cursor()

    cursor_local.execute("SELECT fullname FROM ishchilar WHERE id=? AND role='worker'", (worker_id,))
    res = cursor_local.fetchone()
    
    if res:
        fullname = res[0]
        cursor_local.execute("DELETE FROM ishchilar WHERE id=?", (worker_id,))
        conn_local.commit() 
        await query.answer(f"✅ {fullname} o'chirildi", show_alert=True)
    else:
        await query.answer("❌ Ishchi topilmadi yoki allaqachon o'chirilgan.", show_alert=True)

    cursor_local.execute("SELECT id, fullname FROM ishchilar WHERE role='worker' ORDER BY id")
    workers = cursor_local.fetchall()
    conn_local.close()

    if not workers:
        try:
            await query.message.edit_text("✅ Barcha ishchilar o'chirildi.")
        except Exception:
            pass
        return

    buttons = [[InlineKeyboardButton(f"🗑 {n}", callback_data=f"del_{i}")] for i, n in workers]
    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="back_boss")])
    try:
        await query.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(buttons))
    except Exception:
        await query.message.reply_text("O'chirish uchun ishchini tanlang:", reply_markup=InlineKeyboardMarkup(buttons))

# ================== 15. HISOBOTLAR ==================
async def kunlik(update, context):
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_date = datetime.now().date()
    day_names = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
    day_name = day_names[today_date.weekday()]
    cursor.execute("SELECT fullname, passport FROM ishchilar WHERE role='worker'")
    all_workers = cursor.fetchall()
    cursor.execute("SELECT passport, time FROM davomat WHERE date=? AND confirmed=1", (today_str,))
    present_dict = {p: t for p, t in cursor.fetchall()}
    work_day_note = "✅ Ish kuni" if is_work_day(today_date) else "🔵 Dam olish kuni"
    text = f"🗓 <b>Bugungi davomat — {today_str} ({day_name})</b>\n{work_day_note}\n\n🟢 <b>Kelganlar:</b>\n"
    kelgan_list, kelmagan_list = [], []
    for name, passport in all_workers:
        if passport in present_dict:
            cursor.execute("SELECT late_minutes, jarima FROM moliya WHERE passport=? AND date=?", (passport, today_str))
            mol = cursor.fetchone()
            jarima_str = f" | ⏱ {mol[0]} daq. | 💸 {mol[1]:,} so'm" if mol and mol[0] > 0 else ""
            kelgan_list.append(f"• {name} — {present_dict[passport]}{jarima_str}")
        else:
            kelmagan_list.append(f"• {name}")
    text += "\n".join(kelgan_list) if kelgan_list else "Hali yo'q"
    text += "\n\n🔴 <b>Kelmaganlar:</b>\n"
    text += "\n".join(kelmagan_list) if kelmagan_list else "Hammasi kelgan"
    await update.message.reply_text(text, parse_mode="HTML")

async def oylik_davomad(update, context):
    if update.effective_user.id != BOSS_ID:
        return
    month_str = datetime.now().strftime("%Y-%m")
    ish_kunlari = count_work_days_in_month(month_str)
    file_name = update_excel_report()
    await update.message.reply_document(
        open(file_name, "rb"),
        caption=f"📊 Oylik umumiy davomat — {month_str}\n📅 Ish kunlari: {ish_kunlari}\n🔵 = Dam olish / Bayram kuni"
    )
    os.remove(file_name)

async def send_pdf_report(update, context):
    file = pdf_month_report()
    await update.message.reply_document(open(file, "rb"), caption="📄 Oylik PDF hisoboti")
    os.remove(file)

async def send_oylik_maosh_excel(update, context):
    """Boshliq uchun — barcha ishchilarning oylik maoshi."""
    if update.effective_user.id != BOSS_ID:
        await ishchi_oylik_maosh_request(update, context)
        return
    file = export_oylik_maosh_excel()
    month_str = datetime.now().strftime("%Y-%m")
    ish_kunlari = count_work_days_in_month(month_str)
    await update.message.reply_document(
        open(file, "rb"),
        caption=(
            f"💰 Umumiy oylik maosh — {month_str}\n"
            f"📅 Ish kunlari: {ish_kunlari}\n"
            f"⏰ Ish soati: 08:00–18:00 (9 soat/kun)\n"
            f"📋 Ustunlar: F.I.Sh | Nominal | Kelgan | Kelmagan | Yo'qchilik | Jarima | Avans | Sof maosh"
        )
    )
    os.remove(file)

async def ishchi_oylik_maosh_request(update, context):
    passport = context.user_data.get("passport")
    if not passport:
        await update.message.reply_text("❌ Avval ✅ Keldim bo'limidan kirish kodni kiriting.")
        return
    month_str = datetime.now().strftime("%Y-%m")
    hisob = oylik_maosh_hisob(passport, month_str)
    if not hisob:
        await update.message.reply_text("❌ Ma'lumot topilmadi.")
        return
    
    text = (
        f"👤 <b>F.I.Sh:</b> {hisob['fullname']}\n"
        f"📅 <b>Ish kunlari:</b> {hisob['ish_kunlari']} kun\n"
        f"✅ <b>Kelgan kunlar:</b> {hisob['kelgan']} kun\n"
        f"❌ <b>Kelmagan kunlar:</b> {hisob['kelmagan']} kun\n"
        f"⏱ <b>Kechikish jarimalari:</b> {hisob['jarima']:,} so'm ({hisob['late_minutes']} daq.)\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"💰 <b>Oylik maosh: {hisob['sof_maosh']:,} so'm</b>"
    )
    await update.message.reply_text(text, parse_mode="HTML")
    
    file = export_oylik_maosh_excel_for(passport, month_str)
    if file:
        await update.message.reply_document(open(file, "rb"), caption="📊 Shaxsiy oylik maosh hisoboti (Excel)")
        os.remove(file)

async def ishchi_davomat_request(update, context):
    passport = context.user_data.get("passport")
    if not passport:
        await update.message.reply_text("❌ Avval ✅ Keldim bo'limidan kirish kodni kiriting.")
        return
    file = export_worker_davomat_excel(passport)
    if file:
        await update.message.reply_document(
            open(file, "rb"),
            caption="📋 Sizning shaxsiy davomat hisobotingiz\n🔵 = Dam olish / Bayram kuni"
        )
        os.remove(file)

# ================== 16. GLOBAL NAVIGATSIYA ==================
async def global_orqaga(update, context):
    if "role" in context.user_data:
        del context.user_data["role"]
    kb = ReplyKeyboardMarkup([["👔 Boshliq"], ["👷 Ishchi"]], resize_keyboard=True)
    await update.message.reply_text("Bosh sahifa. Kim sifatida kirasiz?", reply_markup=kb)

async def universal_text_handler(update, context):
    text = update.message.text
    role = context.user_data.get("role")
    if text == "⬅️ Orqaga":
        await global_orqaga(update, context)
        return
    if role == "ishchi":
        await ishchi_main_menu(update, context)
    elif role == "boshliq":
        await boshliq_menu(update, context)
    else:
        await start(update, context)

# ================== 17. FILTR VA CONVERSATION HANDLERS ==================
IGNORE_BTNS = filters.Regex(
    "^(⬅️ Orqaga|📍 Boss lokatsiya|💰 Oylik maosh|📄 PDF chiqarish|📊 Excel chiqarish"
    "|📆 Kunlik|📆 Oylik|➕ Ishchi qo'shish|❌ Ishchi o'chirish|📨 Ish berish|✅ Keldim"
    "|📋 Davomat|📋 Ma'lumot|📩 Topshiriqlar|📌 Eslatmalar)$"
)
INPUT_FILTER = filters.TEXT & ~filters.COMMAND & ~IGNORE_BTNS

async def cancel_conversation(update, context):
    role = context.user_data.get("role")
    if role == "boshliq":
        await boshliq_menu(update, context)
    elif role == "ishchi":
        await ishchi_main_menu(update, context)
    else:
        await global_orqaga(update, context)
    return ConversationHandler.END

conv_add_worker = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^➕ Ishchi qo'shish$"), ishchi_qoshish)],
    states={
        FULLNAME:     [MessageHandler(INPUT_FILTER, ishchi_fullname)],
        PASSPORT:     [MessageHandler(INPUT_FILTER, ishchi_passport)],
        JSHSHIR:      [MessageHandler(INPUT_FILTER, ishchi_jshshir)],
        MAOSH:        [MessageHandler(INPUT_FILTER, ishchi_maosh)],
        PIN_ADD:      [MessageHandler(INPUT_FILTER, ishchi_pin_add)],
        LOCATION:     [MessageHandler(INPUT_FILTER, ishchi_location)],
    },
    fallbacks=[MessageHandler(filters.Regex("^(⬅️ Orqaga|⬅️ Bekor qilish)$"), cancel_conversation)],
    allow_reentry=True
)

conv_keldim = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^✅ Keldim$"), keldim_code_start)],
    states={
        KELDIM_CODE: [MessageHandler(filters.TEXT | filters.LOCATION, check_code_or_location)],
        PIN_STEP:    [MessageHandler(filters.TEXT & ~filters.COMMAND, check_pin_for_keldim)],
    },
    fallbacks=[MessageHandler(filters.Regex("^(⬅️ Orqaga|⬅️ Bekor qilish)$"), cancel_conversation)],
    allow_reentry=True
)

conv_boss_location = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^📍 Boss lokatsiya$"), ask_boss_location)],
    states={BOSS_LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_boss_location)]},
    fallbacks=[MessageHandler(filters.Regex("^(⬅️ Orqaga|⬅️ Bekor qilish)$"), cancel_conversation)]
)

conv_ish_berish = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^📨 Ish berish$"), ish_berish_start)],
    states={
        ISHBERISH_SELECT:  [CallbackQueryHandler(ish_berish_tanlash, pattern="^(ish_|back_boss)")],
        ISHBERISH_MESSAGE: [MessageHandler(filters.TEXT | filters.VOICE, ish_berish_yuborish)],
    },
    fallbacks=[MessageHandler(filters.Regex("^(⬅️ Orqaga|⬅️ Bekor qilish)$"), cancel_conversation)]
)

conv_topshiriqlar = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^📩 Topshiriqlar$"), ishchi_topshiriqlar)],
    states={
        TOPSHIRIQ_JAVOB: [MessageHandler(filters.TEXT | filters.VOICE, ishchi_topshiriq_javob)],
    },
    fallbacks=[MessageHandler(filters.Regex("^⬅️ Orqaga$"), cancel_conversation)],
    allow_reentry=True
)

conv_eslatma_qoshish = ConversationHandler(
    entry_points=[CallbackQueryHandler(handle_eslatma_callback, pattern="^add_eslatma$")],
    states={
        YANGI_ESLATMA: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_yangi_eslatma)],
    },
    fallbacks=[MessageHandler(filters.Regex("^(⬅️ Orqaga|⬅️ Bekor qilish)$"), cancel_conversation)],
    allow_reentry=True
)

# ================== 18. KEEP ALIVE ==================
async def keep_alive():
    url = "https://ish-bot-2.onrender.com"
    while True:
        try:
            async with httpx.AsyncClient() as client:
                await client.get(url)
            print("Ping: Server uyg'oq ushlab turilibdi...")
        except Exception as e:
            print(f"Ping yuborishda xato: {e}")
        await asyncio.sleep(840)

async def post_init(application: Application):
    asyncio.create_task(keep_alive())

# ================== 19. MAIN ==================
def main():
    app = ApplicationBuilder().token(TOKEN).post_init(post_init).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Regex("^(👔 Boshliq|👷 Ishchi)$"), role_handler))

    app.add_handler(conv_ish_berish)
    app.add_handler(conv_add_worker)
    app.add_handler(conv_keldim)
    app.add_handler(conv_boss_location)
    app.add_handler(conv_topshiriqlar)
    app.add_handler(conv_eslatma_qoshish)

    app.add_handler(MessageHandler(filters.Regex("^❌ Ishchi o'chirish$"), ishchi_ochirish))
    app.add_handler(MessageHandler(filters.Regex("^📄 PDF chiqarish$"), send_pdf_report))
    app.add_handler(MessageHandler(filters.Regex("^📊 Excel chiqarish$"), oylik_davomad))
    app.add_handler(MessageHandler(filters.Regex("^💰 Oylik maosh$"), send_oylik_maosh_excel))
    app.add_handler(MessageHandler(filters.Regex("^📆 Kunlik$"), kunlik))
    app.add_handler(MessageHandler(filters.Regex("^📆 Oylik$"), oylik_davomad))
    app.add_handler(MessageHandler(filters.Regex("^📋 Ma'lumot$"), malumot_handler))
    app.add_handler(MessageHandler(filters.Regex("^📌 Eslatmalar$"), eslatmalar_menu))

    app.add_handler(MessageHandler(filters.Regex("^📋 Davomat$"), ishchi_davomat_request))

    app.add_handler(CallbackQueryHandler(handle_eslatma_callback, pattern="^(toggle_eslatma_|del_eslatma_)"))
    app.add_handler(CallbackQueryHandler(delete_worker, pattern="^del_"))
    app.add_handler(CallbackQueryHandler(delete_worker, pattern="^back_boss$"))

    app.add_handler(MessageHandler(filters.Regex("^⬅️ Orqaga$"), global_orqaga))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, universal_text_handler))

    def run_dummy_server():
        try:
            port = int(os.environ.get("PORT", 10000))
            httpd = HTTPServer(("0.0.0.0", port), SimpleHTTPRequestHandler)
            print(f"🌐 Veb-server {port}-portda ishga tushdi")
            httpd.serve_forever()
        except Exception as e:
            print(f"Veb-serverda xato: {e}")

    threading.Thread(target=run_dummy_server, daemon=True).start()
    print("🤖 Bot muvaffaqiyatli ishga tushdi!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()